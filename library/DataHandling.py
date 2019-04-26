from openpyxl import load_workbook


def userName(row_Num, cell_Num, sheet, execlfileName):
	wb = load_workbook(execlfileName, data_only=True)
	# sheet = wb.get_sheet_by_name(sheet)
	sheet = wb[sheet]
	username = sheet.cell(row=row_Num, column=cell_Num).value
	if username is not None:
		return username
	else:
		return ''  # Null Value pass


def password(row_Num, cell_Num, sheet, execlfileName):
	wb = load_workbook(execlfileName, data_only=True)
	# sheet = wb.get_sheet_by_name(sheet)
	sheet = wb[sheet]
	pwd = sheet.cell(row=row_Num, column=cell_Num).value
	if pwd is not None:
		return pwd
	else:
		return ''  # Null Value pass


def excel_handling(row_Num, cell_Num, sheet, execlfileName):
	wb = load_workbook(execlfileName, data_only=True)
	# sheet = wb.get_sheet_by_name(sheet) # Deprecated
	sheet = wb[sheet]
	# gen = str(sheet.cell(row=row_Num, column=cell_Num).value)
	# gen = unicode(sheet.cell(row=row_Num, column=cell_Num).value)
	gen = sheet.cell(row=row_Num, column=cell_Num).value

	if gen is not None:
		gen = str(sheet.cell(row=row_Num, column=cell_Num).value)
		return gen
	elif gen is not None:
		gen = sheet.cell(row=row_Num, column=cell_Num).value
		return gen
	else:
		return ''  # Null Value pass


def cell_length(execlfileName, sheet, name):
	wb = load_workbook(execlfileName, data_only=True)
	sheetName = wb.get_sheet_by_name(sheet)
	cell_len = len(sheetName[name])
	return cell_len
